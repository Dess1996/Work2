from openpyxl import load_workbook
from path_of_excel_files import INDEX_OF_VRP, BEGIN_DATE, AREA

INDEX_VRP = load_workbook(filename=INDEX_OF_VRP)
INDEX_SHEET = INDEX_VRP['Лист1']


def get_years():
    """
    Возвращает заголовки Excel-файла(годы)
    :return: Заголовки
    """
    headers = []
    for cell in INDEX_SHEET[7]:
        headers.append(str(cell.value))
    headers = [char.replace("г.", "") for char in headers]
    headers = list(map(int, headers[1::]))
    begin_years = [year for year in headers if year > BEGIN_DATE]
    return begin_years


def get_areas():
    """
    Получить все значения областей из списка
    :return:
    """
    areas = []
    for cell in INDEX_SHEET['A']:
        if cell.value is None:
            continue
        if 'федеральный' in cell.value:
            continue
        if 'Республика' in cell.value:
            areas.append(cell.value)
        if 'область' in cell.value:
            areas.append(cell.value)
        if 'округ' in cell.value:
            areas.append(cell.value)
    return areas


def get_name_of_indicator():
    indicator_name = ''
    for cell in INDEX_SHEET['A']:
        if cell.value == 'Индексы физического объема валового регионального продукта в 1998-2018гг.':
            indicator_name += cell.value
    return indicator_name


def get_index_vrp(areas):
    """
    Срез значений по годам для одной области
    :param area: область AREA
    :return:  Название области, значения по годам
    """
    area_name = {}
    years = get_years()
    indicator_name = get_name_of_indicator()

    for row in INDEX_SHEET.iter_rows(min_row=1, min_col=1, values_only=True):
        for area in areas:
            if row[0] == area:
                area_name[area] = {}
                values = []
                for val in row[1:]:
                    values.append(val)
                area_name[area] = dict(zip(years, values[len(years) + 1:]))

    res = {indicator_name: area_name}
    return res


if __name__ == '__main__':
    get_years()
    print(get_index_vrp(areas=AREA))
#   print(get_areas())
